def main():
    
    import PyPDF2
    import re
    import pandas as pd
    import openpyxl

    # Caminho da planilha modelo
    caminho_planilha_modelo = 'G:\\Meu Drive\\7. Automação\\OUTRAS AUTOMATIZAÇÕES\\Checklist\\Testes\\Sistema S - SESI SENAI SESC SENAC.xlsx'

    # Aba da planilha modelo
    aba_planilha_modelo = 'Base Dados (Colar Aqui Pgmtos)'

    # Abre o arquivo Excel
    wb = openpyxl.load_workbook(caminho_planilha_modelo)
    ws = wb[aba_planilha_modelo]

    # Pega o cabeçalho da planilha modelo com openpyxl
    itensCabecalho = [cell.value for cell in ws[1]]

    itensCabecalhoFormatado = []

    # Formata o cabeçalho da planilha modelo
    for cont, item in enumerate(itensCabecalho):
        if cont > 0:
            # Pega apenas os 5 primeiros caracteres do item
            item = item[:7]
            itensCabecalhoFormatado.append(item)

    # Lista do resultado
    listaResultado = []

    # Caminho do arquivo PDF
    caminho_arquivo = 'G:\\Meu Drive\\7. Automação\\OUTRAS AUTOMATIZAÇÕES\\Checklist\\Testes\\SEVAN\\comprovantes.pdf'

    # Abre o PDF
    arquivoPDF = open(caminho_arquivo, 'rb')

    # Cria um objeto em PDF
    arquivoPDF = PyPDF2.PdfReader(arquivoPDF)

    # Para cada página do PDF
    for pagina in arquivoPDF.pages:

        # Extrai o texto da página
        texto = pagina.extract_text()
        

        # Capturar as Datas:
        padrao = r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})"
        resultado_REGEX = re.search(padrao, texto)
        apuracao, vencimento = resultado_REGEX.groups()


        # Captura os pagamentos:
        padrao = r"(\d{4})\s+[^\d\n]+?\s+([\d.,]+).*?\n.*?([\d.,]+)"
        resultado_REGEX = re.findall(padrao, texto) # Ele salva em uma tupla, sendo (código, valor, dígito)

        # Para cada Pagamento encontrado
        for codigo, valor, digito in resultado_REGEX:

            # Cria um dicionário para armazenar os dados extraídos
            dadosExtraidos = {}

            # Registra os pagamentos no dicionário
            dadosExtraidos['dtApuracao'] = apuracao
            dadosExtraidos['dtVencimento'] = vencimento
            dadosExtraidos["codigo"] = codigo + "-" + digito
            dadosExtraidos["valor"] = valor

            # Adiciona o dicionário à lista de dados extraídos
            listaResultado.append(dadosExtraidos)

    # Filtra apenas os pagamentos com o código correto
    listaResultadoFiltrada = []
    for pagamento in listaResultado:
        print(pagamento["codigo"])
        if pagamento["codigo"] in itensCabecalhoFormatado:
            listaResultadoFiltrada.append(pagamento)

    # print(listaResultadoFiltrada)


    for pagamento in listaResultadoFiltrada:
        for row in range(2, ws.max_row + 1):  # Começa em 2 para pular o cabeçalho
            data_celula = ws.cell(row=row, column=1).value
            # Se a célula for datetime, formata para string
            if hasattr(data_celula, 'strftime'):
                data_celula_str = data_celula.strftime("%d/%m/%Y")
            else:
                data_celula_str = str(data_celula)
            if pagamento["dtApuracao"] == data_celula_str:
                coluna = itensCabecalhoFormatado.index(pagamento["codigo"]) + 2  # +2 para alinhar com o Excel
                ws.cell(row=row, column=coluna, value=pagamento["valor"])

    try:
        # Salva uma cópia da planilha modelo com os dados preenchidos
        caminho_planilha_modelo = caminho_planilha_modelo.replace('.xlsx', '_preenchida.xlsx')
        wb.save(caminho_planilha_modelo)
    except Exception as e:
        print(f"Erro ao salvar o arquivo: {e}")


if __name__ == "__main__":
    main()
