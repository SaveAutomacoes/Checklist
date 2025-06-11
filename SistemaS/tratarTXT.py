def main(caminho_planilha_modelo, caminho_arquivo, caminho_pasta_cliente, cnpj):
    """
    Processa comprovantes de pagamento do Sistema S extraídos do e-CAC e preenche uma planilha modelo Excel.

    Parâmetros:
    -----------
    caminho_planilha_modelo : str
        Caminho do arquivo Excel modelo que contém o cabeçalho com os códigos de pagamento.
    caminho_arquivo : str
        Caminho do arquivo PDF que contém os comprovantes de pagamento.
    caminho_pasta_cliente : str
        Caminho da pasta onde os arquivos de saída (Excel, logs e PDFs) serão salvos.
    cnpj : str
        CNPJ do cliente, usado para nomear o arquivo Excel salvo e os arquivos de log.

    O que a função faz:
    -------------------
    - Lê um PDF de comprovantes de pagamento do Sistema S.
    - Remove páginas duplicadas (considerando apenas o conteúdo principal, ignorando as 6 últimas linhas de cada página).
    - Extrai datas e valores dos comprovantes usando expressões regulares.
    - Filtra e soma os pagamentos conforme os códigos presentes no cabeçalho da planilha modelo.
    - Preenche a planilha Excel modelo com os valores extraídos.
    - Salva uma cópia da planilha preenchida na pasta do cliente.
    - Gera um log detalhado da execução, incluindo horários, pagamentos processados, filtrados e somados.
    - Salva dois PDFs: um com as páginas consideradas (únicas) e outro com as páginas duplicadas, ambos na subpasta "logs" do cliente.

    Retorno:
    --------
    str
        Caminho do arquivo Excel preenchido salvo na pasta do cliente.
    """
    import PyPDF2
    import re
    import openpyxl
    import os
    from datetime import datetime
    from time import time

    # Define o início
    tempoInicio = time()

    # Aba da planilha modelo onde serão colados os pagamento
    aba_planilha_modelo = 'Base Dados (Colar Aqui Pgmtos)'

    # Define o Dia e Hora atual sem os microssegundos
    hoje = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


    # Cria a pasta "logs" dentro da pasta do cliente, se não existir
    pasta_logs = os.path.join(caminho_pasta_cliente, "logs")
    os.makedirs(pasta_logs, exist_ok=True)

    # Caminho do arquivo de log
    caminho_log = os.path.join(pasta_logs, f"log_{cnpj}_{hoje}.txt")

    logs = []

    def log(msg):
        logs.append(msg)

    # Abre o arquivo Excel
    wb = openpyxl.load_workbook(caminho_planilha_modelo)
    ws = wb[aba_planilha_modelo]

    # Pega o cabeçalho da planilha modelo com openpyxl
    itensCabecalho = [cell.value for cell in ws[1]]

    # Lista para armazenar os itens do cabeçalho formatados
    itensCabecalhoFormatado = [] 

    # Formata o cabeçalho da planilha modelo
    for cont, item in enumerate(itensCabecalho):
        if cont > 0:
            # Pega apenas os 5 primeiros caracteres do item
            item = item[:7]
            itensCabecalhoFormatado.append(item)

    # Lista do resultado
    listaResultado = []

    # Abre o PDF
    arquivoPDF = open(caminho_arquivo, 'rb')

    # Cria um objeto em PDF
    arquivoPDF = PyPDF2.PdfReader(arquivoPDF)

    # Essas variáveis serão usadas para evitar duplicação de páginas, que podem ocorrer no PDF.
    textos_processados = set()
    paginas_duplicadas = []
    paginas_consideradas = []

    for pagina_num, pagina in enumerate(arquivoPDF.pages):
        # Extrai o texto da página
        texto = pagina.extract_text()
        # Remove as 6 últimas linhas do texto, pois o final sempre muda, e queremos ver se estão duplicados.
        linhas = texto.splitlines()
        texto_sem_fim = "\n".join(linhas[:-6]) if len(linhas) > 6 else texto

        if texto_sem_fim in textos_processados:  # Verifica se o texto já foi processado
            log(f"Página {pagina_num+1} ignorada (duplicada).")
            paginas_duplicadas.append(pagina_num)
            continue
        # Se o texto não foi processado, adiciona à lista de textos processados
        textos_processados.add(texto_sem_fim)
        paginas_consideradas.append(pagina_num)
        log(f"\n--- Texto extraído da página {pagina_num+1} ---\n{texto}\n")

        # Capturar as Datas:
        padrao = r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})"
        resultado_REGEX = re.search(padrao, texto)
        apuracao, vencimento = resultado_REGEX.groups()


        # Captura os pagamentos:
        padrao = r"(\d{4})\s+[^\d\n]+?\s+([\d.,]+).*?\n.*?([\d.,]+)"
        resultado_REGEX = re.findall(padrao, texto) # Ele salva em uma tupla, sendo (código, valor, dígito)
        log(f"Pagamentos captados pelo REGEX (página {pagina_num+1}): {resultado_REGEX}")

        # Para cada Pagamento encontrado
        for codigo, valor, digito in resultado_REGEX:

            # Cria um dicionário para armazenar os dados extraídos
            dadosExtraidos = {}

            # Registra os pagamentos no dicionário
            dadosExtraidos['dtApuracao'] = apuracao
            dadosExtraidos['dtVencimento'] = vencimento
            dadosExtraidos["codigo"] = codigo + "-" + digito
            dadosExtraidos["valor"] = valor

            # Adiciona o dicionário à lista de dados extraídos
            listaResultado.append(dadosExtraidos)

    # Filtra apenas os pagamentos com o código correto
    listaResultadoFiltrada = []
    for pagamento in listaResultado:
        if pagamento["codigo"] in itensCabecalhoFormatado:
            listaResultadoFiltrada.append(pagamento)
    log(f"Pagamentos filtrados pelo código: {listaResultadoFiltrada}")

    # Se tiverem pagamentos com a mesma data e código, unifica os valores
    pagamentos_unificados = {}

    for pagamento in listaResultadoFiltrada:
        chave = (pagamento["dtApuracao"], pagamento["codigo"])
        valor_novo = float(pagamento["valor"].replace(".", "").replace(',', '.'))
        if chave not in pagamentos_unificados:
            pagamentos_unificados[chave] = valor_novo
            log(f"Pagamento adicionado na planilha: {chave} valor: {valor_novo}")
        else:
            soma = pagamentos_unificados[chave] + valor_novo
            pagamentos_unificados[chave] = round(soma, 2)
            log(f"Pagamento SOMADO na planilha: {chave} valor anterior: {pagamentos_unificados[chave] - valor_novo} + valor novo: {valor_novo} = {pagamentos_unificados[chave]}")

    for (dtApuracao, codigo), valor in pagamentos_unificados.items():
        for row in range(2, ws.max_row + 1):  # Começa em 2 para pular o cabeçalho
            data_celula = ws.cell(row=row, column=1).value
            # Se a célula for datetime, formata para string
            if hasattr(data_celula, 'strftime'):
                data_celula_str = data_celula.strftime("%d/%m/%Y")
            else:
                data_celula_str = str(data_celula)
            if dtApuracao == data_celula_str:
                coluna = itensCabecalhoFormatado.index(codigo) + 2  # +2 para alinhar com o Excel
                ws.cell(row=row, column=coluna, value=valor)

    # Resumo por data e código
    log("\n--- RESUMO DE PAGAMENTOS ADICIONADOS ---")
    for (dtApuracao, codigo), valor in pagamentos_unificados.items():
        # Procura todos os pagamentos filtrados para essa data e código
        pagamentos = [
            p for p in listaResultadoFiltrada
            if p["dtApuracao"] == dtApuracao and p["codigo"] == codigo
        ]
        if len(pagamentos) == 1:
            log(f"Data: {dtApuracao} | Código: {codigo} | Valor adicionado: {pagamentos[0]['valor']}")
        else:
            valores = [p['valor'] for p in pagamentos]
            log(f"Data: {dtApuracao} | Código: {codigo} | Valores somados: {valores} | Soma final: {valor}")

    try:
        # Salva uma cópia da planilha modelo com os dados preenchidos
        wb.save(caminho_pasta_cliente + f"\Sistema S - {cnpj}.xlsx")
        log("Arquivo Excel salvo com sucesso.")
    except Exception as e:
        log(f"Erro ao salvar o arquivo: {e}")

    # Define o fim
    tempoFim = time()

    # Formata horários para o log
    from datetime import timedelta
    hora_inicio = datetime.fromtimestamp(tempoInicio).strftime("%H:%M:%S")
    hora_fim = datetime.fromtimestamp(tempoFim).strftime("%H:%M:%S")
    duracao = str(timedelta(seconds=int(tempoFim - tempoInicio)))

    # Salva o log no final, adicionando cabeçalho e demais informações
    with open(caminho_log, "w", encoding="utf-8") as f:
        f.write(f"Log de execução - {hoje}\n")
        f.write(f"CNPJ: {cnpj}\n")
        f.write(f"Caminho da planilha modelo: {caminho_planilha_modelo}\n")
        f.write(f"Caminho do arquivo PDF: {caminho_arquivo}\n")
        f.write(f"Caminho da pasta do cliente: {caminho_pasta_cliente}\n\n")
        f.write(f"Início da Execução: {hora_inicio}\n")
        f.write(f"Fim da Execução: {hora_fim}\n")
        f.write(f"Duração da Execução: {duracao}\n")
        f.write("\n--- Início do Log ---\n")

        for linha in logs:
            f.write(str(linha) + "\n")

    # Após o processamento, salve os PDFs separados
    import PyPDF2

    # Caminhos dos PDFs de comprovantes
    caminho_duplicadas = os.path.join(pasta_logs, "Comprovantes de Pagamento Duplicados.pdf")
    caminho_consideradas = os.path.join(pasta_logs, "Comprovantes de Pagamento Considerados.pdf")


    # Salva as páginas duplicadas
    if paginas_duplicadas:
        writer_dup = PyPDF2.PdfWriter()
        with open(caminho_arquivo, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for idx in paginas_duplicadas:
                writer_dup.add_page(reader.pages[idx])
            with open(caminho_duplicadas, "wb") as out_pdf:
                writer_dup.write(out_pdf)
        log(f"{len(paginas_duplicadas)} página(s) duplicada(s) salva(s) em: {caminho_duplicadas}")
    else:
        log("Nenhuma página duplicada encontrada.")

    # Salva as páginas consideradas (únicas)
    if paginas_consideradas:
        writer_cons = PyPDF2.PdfWriter()
        with open(caminho_arquivo, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for idx in paginas_consideradas:
                writer_cons.add_page(reader.pages[idx])
            with open(caminho_consideradas, "wb") as out_pdf:
                writer_cons.write(out_pdf)
        log(f"{len(paginas_consideradas)} página(s) considerada(s) salva(s) em: {caminho_consideradas}")
    else:
        log("Nenhuma página considerada encontrada.")

    return (caminho_pasta_cliente + f"\Sistema S - {cnpj}.xlsx")


# Para Testes

if __name__ == "__main__":
    main(
        caminho_planilha_modelo='G:\\Meu Drive\\7. Automação\\[NÃO EXCLUIR] Documentos-Base\\Modelo Sistema S - SESI SENAI SESC SENAC.xlsx',
        caminho_arquivo='G:\\Meu Drive\\7. Automação\\OUTRAS AUTOMATIZAÇÕES\\Checklist\\Testes\\Teste Checklist\\SEVAN\\Comprovantes de Pagamento.pdf',
        caminho_pasta_cliente='G:\\Meu Drive\\7. Automação\\OUTRAS AUTOMATIZAÇÕES\\Checklist\\Testes\\Teste Checklist\\SEVAN',
        cnpj='3904323000109'
    )